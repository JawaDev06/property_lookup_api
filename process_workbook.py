from __future__ import annotations

import json
import os
import re
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Optional

from dotenv import load_dotenv
from openpyxl import load_workbook

from clients.base import PropertyLookupClient
from clients.smarty_property_data import SmartyPropertyDataClient
from column_map import COLUMNS, DEFAULT_START_ROW
from models import PropertyData

load_dotenv()


@dataclass
class ProcessStats:
    rows_seen: int = 0
    rows_with_address: int = 0
    rows_updated: int = 0
    rows_not_found: int = 0
    errors: int = 0


def make_client() -> PropertyLookupClient:
    # This package intentionally uses Smarty only. No county/GIS/assessor websites are called.
    return SmartyPropertyDataClient()


def _env_col(name: str, default: str) -> str:
    """Return an output column.

    Defaults the extra columns on so you do not have to edit .env.
    To disable any extra field, set its env var to DISABLED, NONE, or OFF.
    """
    raw = os.getenv(name)
    value = (raw if raw is not None else default).strip().upper()
    if value in {"", "DISABLED", "NONE", "OFF", "FALSE", "0"}:
        return ""
    return value


def _optional_extra_columns() -> dict[str, tuple[str, object, str, str]]:
    """Write optional headers when blank.

    v9 keeps the compact garage value in R and restores last sold date/amount
    in U/V while still removing the older extra output columns W:AD by default.
    """
    return {
        COLUMNS.garage: ("garage_display", None, "General", "Garage A/D SqFt"),
        COLUMNS.last_sold_date: ("last_sold_date", None, "m/d/yyyy", "Last Sold Date"),
        COLUMNS.last_sold_amount: ("last_sold_amount", None, "$#,##0", "Last Sold Amount"),
    }


def _garage_display(data: PropertyData):
    """Return a compact garage value for column R.

    Examples:
      A 420  = attached garage, 420 sqft
      D 240  = detached garage, 240 sqft
      A      = attached garage, sqft not returned
      420    = garage sqft returned, attached/detached unknown
    """
    code = (data.garage_code or "").strip().upper() or None
    sqft = data.garage_sqft
    if code and sqft:
        return f"{code} {sqft}"
    if code:
        return code
    if sqft:
        return sqft
    return None


def _remove_columns_u_ad(ws) -> None:
    """Delete old extra columns W:AD by default, while preserving U/V sales data.

    New preferred variable: SMARTY_DELETE_COLUMNS_W_AD.
    Backward-compatible variable: SMARTY_DELETE_COLUMNS_U_AD. If either is set
    false, the worksheet layout is left untouched.
    """
    should_delete = _env_bool(
        "SMARTY_DELETE_COLUMNS_W_AD",
        _env_bool("SMARTY_DELETE_COLUMNS_U_AD", True),
    )
    if not should_delete:
        return
    # W = 23 and AD = 30, so delete 8 columns. This keeps U and V.
    ws.delete_cols(23, 8)


def _env_bool(name: str, default: bool = False) -> bool:
    raw = os.getenv(name)
    if raw is None or raw.strip() == "":
        return default
    return raw.strip().lower() in {"1", "true", "yes", "y", "on"}


def _write_extra_headers(ws) -> None:
    if not _env_bool("SMARTY_WRITE_EXTRA_HEADERS", True):
        return
    header_row = int(os.getenv("SMARTY_EXTRA_HEADER_ROW", "2"))
    for col, (_field_name, _placeholder, _num_fmt, label) in _optional_extra_columns().items():
        if not col:
            continue
        cell = ws[f"{col}{header_row}"]
        if cell.value in (None, ""):
            cell.value = label


def _safe_filename_piece(value: str, max_len: int = 60) -> str:
    text = re.sub(r"[^A-Za-z0-9._-]+", "_", value.strip())
    return text[:max_len].strip("._-") or "property"


def _dump_raw_json(row: int, address: str, data: PropertyData) -> None:
    raw_dir = os.getenv("SMARTY_RAW_JSON_DIR", "").strip()
    if not raw_dir or not data.raw:
        return
    out_dir = Path(raw_dir).expanduser()
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"row_{row}_{_safe_filename_piece(address)}.json"
    out_path.write_text(json.dumps(data.raw, indent=2, default=str), encoding="utf-8")


def _write_property_data(ws, row: int, data: PropertyData) -> None:
    values = {
        COLUMNS.latest_assessed_value: data.latest_assessed_value,
        COLUMNS.square_feet: data.square_feet,
        COLUMNS.year_built: data.year_built,
        COLUMNS.beds: data.beds,
        COLUMNS.baths: data.baths,
        COLUMNS.garage: _garage_display(data),
        COLUMNS.last_sold_date: data.last_sold_date,
        COLUMNS.last_sold_amount: data.last_sold_amount,
    }

    for col, value in values.items():
        if value is not None:
            ws[f"{col}{row}"] = float(value) if hasattr(value, "as_tuple") else value

    ws[f"{COLUMNS.latest_assessed_value}{row}"].number_format = '$#,##0'
    ws[f"{COLUMNS.square_feet}{row}"].number_format = '#,##0'
    ws[f"{COLUMNS.year_built}{row}"].number_format = '0'
    ws[f"{COLUMNS.beds}{row}"].number_format = '0.#'
    ws[f"{COLUMNS.baths}{row}"].number_format = '0.0'
    ws[f"{COLUMNS.garage}{row}"].number_format = 'General'
    ws[f"{COLUMNS.last_sold_date}{row}"].number_format = 'm/d/yyyy'
    ws[f"{COLUMNS.last_sold_amount}{row}"].number_format = '$#,##0'



def process_workbook(
    input_path: str | Path,
    output_path: str | Path,
    client: Optional[PropertyLookupClient] = None,
    sheet_name: Optional[str] = None,
    start_row: int = DEFAULT_START_ROW,
    city_column: Optional[str] = None,
    verbose: bool = True,
) -> ProcessStats:
    input_path = Path(input_path)
    output_path = Path(output_path)
    client = client or make_client()

    wb = load_workbook(input_path)
    ws = wb[sheet_name] if sheet_name else wb.active
    _write_extra_headers(ws)
    stats = ProcessStats()

    try:
        for row in range(start_row, ws.max_row + 1):
            stats.rows_seen += 1
            address = ws[f"{COLUMNS.address}{row}"].value
            if not address or not str(address).strip():
                continue
            stats.rows_with_address += 1
            city_value = ws[f"{city_column}{row}"].value if city_column else None
            city = str(city_value).strip() if city_value else None
            address_text = str(address).strip()
            if verbose:
                print(f"Row {row}: Smarty lookup for {address_text}{', ' + city if city else ''}", flush=True)
            try:
                data = client.lookup(address_text, city=city)
                if data is None:
                    stats.rows_not_found += 1
                    if verbose:
                        print(f"Row {row}: no match", flush=True)
                    continue
                _write_property_data(ws, row, data)
                _dump_raw_json(row, address_text, data)
                stats.rows_updated += 1
                if verbose:
                    print(f"Row {row}: updated", flush=True)
            except Exception as exc:
                stats.errors += 1
                if verbose:
                    print(f"Row {row}: ERROR {exc}", flush=True)
    finally:
        close = getattr(client, "close", None)
        if callable(close):
            close()

    _remove_columns_u_ad(ws)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return stats


if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="Fill weekly workbook property columns using Smarty US Property Data only. v9 writes garage to R, restores last sold date/amount in U/V, removes W:AD, and does not add formatting/comments.")
    parser.add_argument("input", help="Path to input .xlsx")
    parser.add_argument("output", help="Path for updated .xlsx")
    parser.add_argument("--sheet", default=None, help="Worksheet name. Defaults to active sheet.")
    parser.add_argument("--start-row", default=DEFAULT_START_ROW, type=int)
    parser.add_argument("--city-column", default=os.getenv("SMARTY_CITY_COLUMN"), help="Optional Excel column containing city, e.g. C.")
    parser.add_argument("--quiet", action="store_true", help="Suppress row-by-row progress messages.")
    args = parser.parse_args()
    stats = process_workbook(
        args.input,
        args.output,
        sheet_name=args.sheet,
        start_row=args.start_row,
        city_column=args.city_column,
        verbose=not args.quiet,
    )
    print(asdict(stats))
