from __future__ import annotations
import json, shutil
from datetime import datetime
from pathlib import Path
from typing import Iterable
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from property_scraper.common import clean_text, is_probable_address, detect_city, parse_address
from property_scraper.scrapers import scrape_property

COL_ADDRESS = 2   # B
COL_VALUE = 6     # F
COL_SQFT = 14     # N
COL_YEAR = 15     # O
COL_BEDS = 16     # P
COL_BATHS = 17    # Q
COL_STATUS = 18   # R
COL_SOURCE = 19   # S
COL_NOTES = 20    # T
COL_NORMALIZED = 21 # U

HEADERS = {
    COL_VALUE: "Assessed Value",
    COL_SQFT: "Square Feet",
    COL_YEAR: "Year Built",
    COL_BEDS: "Beds",
    COL_BATHS: "Baths",
    COL_STATUS: "Scrape Status",
    COL_SOURCE: "Source URL",
    COL_NOTES: "Scrape Notes",
    COL_NORMALIZED: "Normalized Address",
}


def ensure_headers(ws: Worksheet):
    for col, text in HEADERS.items():
        if not clean_text(ws.cell(row=2, column=col).value):
            ws.cell(row=2, column=col).value = text


def write_result(ws: Worksheet, row: int, result, overwrite: bool):
    mapping = [
        (COL_VALUE, result.assessed_value),
        (COL_SQFT, result.sqft),
        (COL_YEAR, result.year_built),
        (COL_BEDS, result.beds),
        (COL_BATHS, result.baths),
        (COL_STATUS, result.status),
        (COL_SOURCE, result.source_url),
        (COL_NOTES, result.notes),
        (COL_NORMALIZED, result.normalized_address),
    ]
    for col, value in mapping:
        if overwrite or not clean_text(ws.cell(row=row, column=col).value):
            ws.cell(row=row, column=col).value = value


def process_workbook(path: Path, output_dir: Path, processed_dir: Path, logs_dir: Path, *, overwrite: bool = False, headed: bool = False, dry_run: bool = False) -> Path:
    wb = load_workbook(path)
    ws = wb.active
    ensure_headers(ws)
    log_rows = []
    for row in range(1, ws.max_row + 1):
        address = clean_text(ws.cell(row=row, column=COL_ADDRESS).value)
        if not is_probable_address(address):
            if address:
                log_rows.append({"row": row, "address": address, "status": "skipped_non_address"})
            continue
        city = detect_city(address)
        parsed = parse_address(address, city)
        result = scrape_property(address, city, headed=headed)
        if not result.city:
            result.city = city
        if not result.normalized_address:
            result.normalized_address = parsed["normalized"]
        write_result(ws, row, result, overwrite=overwrite)
        d = result.to_dict()
        d.update({"row": row, "address": address})
        log_rows.append(d)

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_name = path.stem + "_UPDATED.xlsx"
    out_path = output_dir / out_name
    log_path = logs_dir / f"{path.stem}_scrape_log_{ts}.json"
    output_dir.mkdir(parents=True, exist_ok=True)
    processed_dir.mkdir(parents=True, exist_ok=True)
    logs_dir.mkdir(parents=True, exist_ok=True)
    if not dry_run:
        wb.save(out_path)
        with open(log_path, "w", encoding="utf-8") as f:
            json.dump(log_rows, f, indent=2, ensure_ascii=False)
        shutil.move(str(path), str(processed_dir / path.name))
    return out_path


def process_folder(input_dir: Path, output_dir: Path, processed_dir: Path, logs_dir: Path, *, overwrite: bool = False, headed: bool = False, dry_run: bool = False) -> list[Path]:
    input_dir.mkdir(parents=True, exist_ok=True)
    output_dir.mkdir(parents=True, exist_ok=True)
    processed_dir.mkdir(parents=True, exist_ok=True)
    logs_dir.mkdir(parents=True, exist_ok=True)
    results = []
    for path in sorted(input_dir.glob("*.xlsx")):
        if path.name.startswith("~$"):
            continue
        results.append(process_workbook(path, output_dir, processed_dir, logs_dir, overwrite=overwrite, headed=headed, dry_run=dry_run))
    return results
