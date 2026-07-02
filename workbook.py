import json
import shutil
from datetime import datetime
from pathlib import Path
from typing import Dict, Any
from openpyxl import load_workbook
from router import lookup_property
from config import INPUT_DIR, OUTPUT_DIR, PROCESSED_DIR, LOGS_DIR

ADDRESS_COL = "B"
COLS = {
    "assessed_value": "F",
    "sqft": "N",
    "year_built": "O",
    "beds": "P",
    "baths": "Q",
    "status": "R",
    "source_url": "S",
    "notes": "T",
}


def _write(ws, row: int, result, overwrite: bool = False):
    values = result.to_dict()
    for field, col in COLS.items():
        cell = ws[f"{col}{row}"]
        val = values.get(field)
        if overwrite or cell.value in (None, ""):
            cell.value = val


def process_workbook(input_path: Path, output_path: Path | None = None, overwrite: bool = False) -> Dict[str, Any]:
    input_path = Path(input_path)
    if output_path is None:
        output_path = OUTPUT_DIR / f"{input_path.stem}_UPDATED{input_path.suffix}"
    wb = load_workbook(input_path)
    ws = wb.active
    logs = []
    counts = {}
    for row in range(2, ws.max_row + 1):
        raw = ws[f"{ADDRESS_COL}{row}"].value
        if raw is None or str(raw).strip() == "":
            continue
        result = lookup_property(str(raw))
        _write(ws, row, result, overwrite=overwrite)
        item = {"row": row, **result.to_dict()}
        logs.append(item)
        counts[result.status] = counts.get(result.status, 0) + 1
    # diagnostics sheet
    if "Diagnostics" in wb.sheetnames:
        del wb["Diagnostics"]
    diag = wb.create_sheet("Diagnostics")
    headers = ["row", "address", "city", "status", "assessed_value", "sqft", "year_built", "beds", "baths", "source_url", "notes"]
    diag.append(headers)
    for item in logs:
        diag.append([item.get(h) for h in headers])
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    log_path = LOGS_DIR / f"{input_path.stem}_scrape_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
    with open(log_path, "w", encoding="utf-8") as f:
        json.dump(logs, f, indent=2)
    return {"input": str(input_path), "output": str(output_path), "log": str(log_path), "rows": len(logs), "counts": counts}


def process_input_folder(overwrite: bool = False, move_processed: bool = True) -> Dict[str, Any]:
    summaries = []
    for path in sorted(INPUT_DIR.glob("*.xlsx")):
        if path.name.startswith("~$"):
            continue
        summary = process_workbook(path, overwrite=overwrite)
        summaries.append(summary)
        if move_processed:
            dest = PROCESSED_DIR / path.name
            if dest.exists():
                dest = PROCESSED_DIR / f"{path.stem}_{datetime.now().strftime('%Y%m%d_%H%M%S')}{path.suffix}"
            shutil.move(str(path), dest)
            summary["processed_original"] = str(dest)
    return {"processed_files": len(summaries), "files": summaries}
